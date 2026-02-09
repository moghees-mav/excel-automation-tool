import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

df = pd.read_excel('data/input.xlsx')

df = df.dropna(how='all')
df = df.drop_duplicates()

df.columns = df.columns.str.lower().str.replace(' ', '_').str.replace('[^a-z0-9_]', '', regex=True)

total_rows = len(df)
total_cols = len(df.columns)

numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
numeric_sums = {col: df[col].sum() for col in numeric_cols}

summary_records = [
    {'Metric': 'Total Rows', 'Value': total_rows},
    {'Metric': 'Total Columns', 'Value': total_cols}
]
summary_records.extend([{'Metric': f'Sum of {col}', 'Value': numeric_sums[col]} for col in numeric_cols])

summary_df = pd.DataFrame(summary_records)

with pd.ExcelWriter('data/output.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='CleanedData', index=False)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

wb = load_workbook('data/output.xlsx')

for sheet_name in ['CleanedData', 'Summary']:
    ws = wb[sheet_name]
    
    for col_idx, col in enumerate(ws.columns, 1):
        max_width = 0
        col_letter = get_column_letter(col_idx)
        
        for cell in col:
            try:
                cell_width = len(str(cell.value))
                max_width = max(max_width, cell_width)
            except:
                pass
        
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)
    
    for row in ws.rows:
        max_height = 15
        for cell in row:
            if cell.value:
                line_count = len(str(cell.value).split('\n'))
                max_height = max(max_height, line_count * 15)
        
        ws.row_dimensions[row[0].row].height = max_height

wb.save('data/output.xlsx')
